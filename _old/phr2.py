from openpyxl import load_workbook
import random as r
import pyphen
import inspect
from PIL import Image, ImageDraw, ImageFont

if not hasattr(inspect, 'getargspec'):
    inspect.getargspec = inspect.getfullargspec

def phr():

    wb = load_workbook('phrases.xlsx')
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    ws = wb.active

    row =  r.randint(1,row_count)

    cb = ws['B'][row].value
    cd = ws['D'][row].value
    phrase = (cb + ' ' + cd)

    font = ImageFont.truetype("ds_moster.ttf", 100)
    font2 = ImageFont.truetype("arial.ttf", 40)
    im = Image.new("RGB", (1000, 1000), "black")
    d = ImageDraw.Draw(im)
    # d.line(((0, 500), (1000, 500)), "black")
    # d.line(((500, 0), (500, 1000)), "black")
    d.text((500, 950), phrase, fill="gray", anchor="ms", font=font2)

    # Создаем объект Pyphen для русского языка
    dic = pyphen.Pyphen(lang='ru')
    glas   = "АЕЁИОУЫЭЮЯ"
    glas_s = "аеёиоуыэюя"

    # cb = 'Подложить'
    # cd = 'свинью'
    word1 = cb.upper()
    word2 = cd.upper()
    
    # Разбиваем слово на слоги
    syllables1 = dic.inserted(word1, hyphen='•')
    syllables2 = dic.inserted(word2, hyphen='•')

    # Перемешиваем слоги
    syllables_list = syllables1.split('•')
    syllables_list_2 = syllables2.split('•')

    syllables_list.extend(syllables_list_2)

    for i in syllables_list:
        if (str(i[0]) in glas and len(i)!=2):
            syllables_list[syllables_list.index(i)] = i[0]
            syllables_list.append(i[1:])
        elif (str(i[-1]) in glas and str(i[-2]) in glas):
            syllables_list[syllables_list.index(i)] = i[-1]
            syllables_list.append(i[0:-1])


    r.shuffle(syllables_list)
    shuffled_word = ' '.join(syllables_list)

    sep = int(len(syllables_list) / 2)

    first = ''.join(syllables_list[:sep])
    second = ''.join(syllables_list[sep:])
    nphr = first + ' ' + second

    #colorize slog
    colors = ['Red','Lime','Blue','Orange','Purple','Yellow','Cyan','Pink','Teal','Olive','Green']
    lets = 75
    lp1 = len(first)*lets/2
    lp2 = len(second)*lets/2
    x = 500 - lp1
    y = 350

    for i, letter in enumerate(syllables_list):
        color = colors[i % len(colors)]
        if (i<sep):
            d.text((x, y), letter, fill=color, font=font)
            x += len(letter)*lets
        if (i+1==sep):
            x = 500 - lp2
            # d.text((x+len(letter)*50-30, y+150), '_', fill='black', font=font)
        elif (i+1>sep):
            d.text((x, y+150), letter, fill=color, font=font)
            x += len(letter)*lets

    # im.show()
    im.save("phrase.jpg")

    return phrase, nphr

# phr()